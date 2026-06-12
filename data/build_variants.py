"""
Convert euro polo (finalised).xlsx → product-data.json + product-data-embed.js
with Shopee-style variant structure.

Usage:
  py data/build_variants.py            # process spreadsheet only
  py data/build_variants.py --upload   # also migrate all images to Cloudinary
                                        # requires: pip install requests
"""
import openpyxl, json, re, os, sys, hashlib, time
from collections import defaultdict, Counter

XLSX          = r'C:\Users\User\Desktop\euro polo.web\europolo-website\finaliseddd.xlsx'
COLOUR_XLSX   = r'C:\Users\User\Desktop\euro polo.web\europolo-website\Euro_Polo_Colour_Images.xlsx'
JSON_OUT      = r'C:\Users\User\Desktop\euro polo.web\europolo-website\data\product-data.json'
JS_OUT        = r'C:\Users\User\Desktop\euro polo.web\europolo-website\js\product-data-embed.js'
CACHE_OUT     = r'C:\Users\User\Desktop\euro polo.web\europolo-website\data\cloudinary-cache.json'
ENV_FILE      = r'C:\Users\User\Desktop\euro polo.web\europolo-website\.env'

UPLOAD_MODE = '--upload' in sys.argv

# ── .env loader (no external library) ─────────────────────────────────────────
def load_env(path):
    env = {}
    try:
        with open(path, encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith('#') or '=' not in line:
                    continue
                k, _, v = line.partition('=')
                env[k.strip()] = v.strip()
    except FileNotFoundError:
        pass
    return env

# ── Cloudinary helpers (only initialised in --upload mode) ────────────────────
if UPLOAD_MODE:
    _env = load_env(ENV_FILE)
    CLOUD_NAME = _env.get('CLOUDINARY_CLOUD_NAME') or os.environ.get('CLOUDINARY_CLOUD_NAME', '')
    API_KEY    = _env.get('CLOUDINARY_API_KEY')    or os.environ.get('CLOUDINARY_API_KEY', '')
    API_SECRET = _env.get('CLOUDINARY_API_SECRET') or os.environ.get('CLOUDINARY_API_SECRET', '')

    missing = [k for k, v in [('CLOUDINARY_CLOUD_NAME', CLOUD_NAME),
                               ('CLOUDINARY_API_KEY',    API_KEY),
                               ('CLOUDINARY_API_SECRET', API_SECRET)] if not v]
    if missing:
        print(f'ERROR: Missing in .env: {", ".join(missing)}')
        sys.exit(1)

    try:
        import requests as _req
    except ImportError:
        print('ERROR: pip install requests   (required for --upload)')
        sys.exit(1)

    # Load URL-to-Cloudinary cache (skip already-uploaded images on re-runs)
    try:
        with open(CACHE_OUT, encoding='utf-8') as _cf:
            _cache = json.load(_cf)
    except (FileNotFoundError, json.JSONDecodeError):
        _cache = {}

    def _safe_id(s):
        """Sanitise a string to a valid Cloudinary public_id segment."""
        return re.sub(r'[^a-zA-Z0-9_-]', '_', str(s))

    def _is_cloudinary(url):
        return bool(url and 'res.cloudinary.com' in str(url))

    def _upload(original_url, public_id):
        """Download image from original_url and upload to Cloudinary (signed).
        Returns the Cloudinary secure_url, or original_url on failure."""
        if not original_url or _is_cloudinary(original_url):
            return original_url
        if original_url in _cache:
            return _cache[original_url]

        # Download from Shopee CDN
        try:
            r = _req.get(
                original_url,
                headers={'User-Agent': 'Mozilla/5.0',
                         'Referer':     'https://shopee.com.my/'},
                timeout=30,
            )
            r.raise_for_status()
        except Exception as e:
            print(f'\n  [WARN] download failed ({public_id}): {e}')
            return original_url

        # Build SHA-1 signature (params sorted alphabetically, excluding api_key/file)
        ts  = str(int(time.time()))
        sig = hashlib.sha1(
            (f'public_id={public_id}&timestamp={ts}' + API_SECRET).encode()
        ).hexdigest()

        # Signed upload (multipart via requests)
        try:
            resp = _req.post(
                f'https://api.cloudinary.com/v1_1/{CLOUD_NAME}/image/upload',
                data={
                    'api_key':   API_KEY,
                    'timestamp': ts,
                    'signature': sig,
                    'public_id': public_id,
                },
                files={'file': ('img', r.content,
                                r.headers.get('content-type', 'image/jpeg'))},
                timeout=60,
            )
            if resp.ok:
                url = resp.json()['secure_url']
                _cache[original_url] = url
                return url
            else:
                print(f'\n  [WARN] upload HTTP {resp.status_code} ({public_id}): {resp.text[:150]}')
                return original_url
        except Exception as e:
            print(f'\n  [WARN] upload error ({public_id}): {e}')
            return original_url

    def _save_cache():
        with open(CACHE_OUT, 'w', encoding='utf-8') as f:
            json.dump(_cache, f, indent=2)

# ── Read spreadsheet ───────────────────────────────────────────────────────────
wb   = openpyxl.load_workbook(XLSX)
ws_p = wb['Products']
ws_v = wb['Variants']

def read_sheet(ws):
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {h: ws.cell(r, c).value for c, h in enumerate(headers, 1)}
        rows.append(row)
    return rows

products_rows = [r for r in read_sheet(ws_p) if r.get('product_id')]
variants_rows = [r for r in read_sheet(ws_v) if r.get('product_id')]

# ── Helpers ────────────────────────────────────────────────────────────────────
def clean(v):
    if v is None: return None
    s = str(v).strip()
    return None if s.lower() in ('nan', 'none', '') else s

def parse_gallery(raw):
    if not raw: return []
    return [u.strip() for u in re.split(r'[\r\n]+', str(raw))
            if u.strip() and u.strip().lower() not in ('nan', 'none')]

def fmt_price(lo, hi):
    if abs(lo - hi) < 0.005:
        return 'RM {:,.2f}'.format(lo)
    return 'RM {:,.2f} – RM {:,.2f}'.format(lo, hi)

# ── Category mapping ───────────────────────────────────────────────────────────
def map_cat(product_type, category, name):
    pt  = str(product_type or '').lower()
    cat = str(category     or '').lower()
    nm  = str(name         or '').lower()

    if 'luggage' in pt or 'luggage' in cat:
        return 'luggage', 'luggage'
    if 'belt' in pt:
        return 'belts', 'belts'
    if 'gift set' in pt:
        return 'wallets', 'bifold-trifold'
    if 'wallet' in pt or 'card holder' in pt or 'card' in pt:
        if 'card' in cat or 'card holder' in nm:
            return 'wallets', 'card-holders'
        if 'long' in nm or 'long wallet' in cat:
            return 'wallets', 'long-wallets'
        return 'wallets', 'bifold-trifold'
    if 'bag' in pt or 'bag' in cat:
        if 'crossbody' in cat or 'shoulder' in cat or 'sling' in cat:
            return 'bags', 'crossbody'
        if 'backpack' in cat:
            return 'bags', 'backpacks'
        if 'laptop' in cat:
            return 'bags', 'laptop-bags'
        if 'waist' in cat or 'chest' in cat:
            return 'bags', 'waist-chest'
        return 'bags', 'crossbody'
    return 'bags', 'crossbody'

# ── Group variants by product ──────────────────────────────────────────────────
variants_by_pid = defaultdict(list)
for row in variants_rows:
    pid = str(row['product_id']).strip()
    variants_by_pid[pid].append(row)

# ── Load colour-image map from Euro_Polo_Colour_Images.xlsx ───────────────────
colour_map = defaultdict(dict)   # colour_map[product_id][colour] = image_url
colour_map_total = 0
try:
    wb_c  = openpyxl.load_workbook(COLOUR_XLSX, read_only=True)
    ws_c  = wb_c['Colour Images']
    _rows = list(ws_c.iter_rows(values_only=True))
    for _row in _rows[1:]:                          # skip header
        _pid, _name, _colour, _filename, _url = _row
        if _pid and _colour and _url:
            colour_map[str(_pid).strip()][str(_colour).strip()] = str(_url).strip()
            colour_map_total += 1
    wb_c.close()
    print(f'Colour map loaded: {colour_map_total} entries across {len(colour_map)} products')
except FileNotFoundError:
    print(f'  [WARN] Euro_Polo_Colour_Images.xlsx not found — variant images will use fallback only')
except Exception as _e:
    print(f'  [WARN] Could not load colour map: {_e}')

colour_map_hits    = 0
colour_map_misses  = []

# ── Build output ───────────────────────────────────────────────────────────────
output  = []
skipped = []

for p_row in products_rows:
    pid   = str(p_row['product_id']).strip()
    pvars = variants_by_pid.get(pid, [])
    s_cat, s_sub = map_cat(p_row.get('product_type'), p_row.get('category'), p_row.get('product_name'))

    variant_list = []
    sku_seen = set()
    for v in pvars:
        sku = clean(v.get('variant_sku'))
        if not sku:
            o1  = str(v.get('option1_value') or 'X').strip()[:6].replace(' ', '')
            o2  = str(v.get('option2_value') or '').strip()[:4].replace(' ', '')
            sku = pid + '-' + o1 + (('-' + o2) if o2 else '')
            print(f'  [WARN] {pid}: blank SKU → generated "{sku}"')

        if sku in sku_seen:
            sku = sku + '-dup'
        sku_seen.add(sku)

        try:
            price = float(v['price'])
        except (TypeError, ValueError):
            print(f'  [WARN] {pid} {sku}: invalid price "{v.get("price")}", skipping variant')
            continue

        try:
            stock = int(float(str(v['stock']))) if v.get('stock') is not None else 0
        except (TypeError, ValueError):
            stock = 0

        opt1 = clean(v.get('option1_value'))
        opt2 = clean(v.get('option2_value'))
        img  = clean(v.get('variant_image'))

        # Inject colour-specific image from Euro_Polo_Colour_Images.xlsx
        if opt1 and pid in colour_map and opt1 in colour_map[pid]:
            img = colour_map[pid][opt1]
            colour_map_hits += 1
        elif opt1:
            colour_map_misses.append((pid, opt1))
            print(f'  [WARN] No colour image mapping found for {pid} / {opt1}')

        variant_list.append({
            'sku':     sku,
            'option1': opt1,
            'option2': opt2,
            'price':   price,
            'stock':   stock,
            'image':   img,
        })

    if not variant_list:
        skipped.append(pid)
        print(f'  [SKIP] {pid}: no valid variants')
        continue

    prices    = [v['price'] for v in variant_list]
    price_min = min(prices)
    price_max = max(prices)
    in_stock  = any(v['stock'] > 0 for v in variant_list)

    gallery   = parse_gallery(p_row.get('gallery_images'))
    cover_img = clean(p_row.get('cover_image'))
    if not cover_img:
        for v in variant_list:
            if v['image']:
                cover_img = v['image']
                break

    opt1_name = clean(p_row.get('option1_name'))
    opt2_name = clean(p_row.get('option2_name'))

    if opt2_name and opt2_name.lower() in ('option 2', 'option2'):
        print(f'  [WARN] {pid}: option2_name is placeholder "{opt2_name}"')

    output.append({
        'id':          pid,
        'name':        str(p_row.get('product_name', '') or '').strip(),
        'productType': str(p_row.get('product_type',  '') or '').strip(),
        'category':    s_cat,
        'subcategory': s_sub,
        'option1Name': opt1_name,
        'option2Name': opt2_name,
        'description': str(p_row.get('description',   '') or '').strip(),
        'images': {
            'cover':   cover_img,
            'gallery': gallery,
        },
        'variants':     variant_list,
        'priceMin':     price_min,
        'priceMax':     price_max,
        'priceDisplay': fmt_price(price_min, price_max),
        'inStock':      in_stock,
    })

# ── Stats ──────────────────────────────────────────────────────────────────────
total_variants = sum(len(p['variants']) for p in output)
oos_variants   = sum(1 for p in output for v in p['variants'] if v['stock'] == 0)
two_option     = sum(1 for p in output if p['option2Name'])

print()
print('=' * 56)
print(f'Products processed : {len(output)}')
print(f'Products skipped   : {len(skipped)}')
print(f'Total variants     : {total_variants}')
print(f'Out-of-stock vars  : {oos_variants}')
print(f'Two-option products: {two_option}')
print('=' * 56)
cat_counts = Counter(p['category'] for p in output)
for cat, n in sorted(cat_counts.items()):
    print(f'  {cat:<10} {n} products')
print()
print('-- Colour image mapping --------------------------------------')
print(f'  Mappings loaded    : {colour_map_total}')
print(f'  Variants mapped    : {colour_map_hits}')
print(f'  Missing mappings   : {len(colour_map_misses)}')
if colour_map_misses:
    for _pid, _col in colour_map_misses:
        print(f'    MISSING: {_pid} / {_col}')
print()

# ── Cloudinary image migration (--upload only) ─────────────────────────────────
if UPLOAD_MODE:
    # Count images to process (exclude already-Cloudinary and None)
    all_urls = set()
    for p in output:
        if p['images']['cover']   and not _is_cloudinary(p['images']['cover']):
            all_urls.add(p['images']['cover'])
        for g in p['images']['gallery']:
            if g and not _is_cloudinary(g):
                all_urls.add(g)
        for v in p['variants']:
            if v['image'] and not _is_cloudinary(v['image']):
                all_urls.add(v['image'])

    already_cached = sum(1 for u in all_urls if u in _cache)
    to_upload      = len(all_urls) - already_cached

    print(f'Image migration: {len(all_urls)} unique Shopee URLs found')
    print(f'  {already_cached} already cached, {to_upload} new uploads needed')
    print()

    done = 0
    total = 0
    for p in output:
        total += bool(p['images']['cover'])
        total += len(p['images']['gallery'])
        total += sum(1 for v in p['variants'] if v['image'])

    for p in output:
        pid = p['id']

        if p['images']['cover']:
            done += 1
            print(f'  [{done}/{total}] {pid}/cover', end='\r', flush=True)
            p['images']['cover'] = _upload(
                p['images']['cover'], f'europolo/{pid}/cover')

        for i, gurl in enumerate(p['images']['gallery']):
            if gurl:
                done += 1
                print(f'  [{done}/{total}] {pid}/gallery-{i}   ', end='\r', flush=True)
                p['images']['gallery'][i] = _upload(
                    gurl, f'europolo/{pid}/g{i}')

        for v in p['variants']:
            if v['image']:
                done += 1
                sid = _safe_id(v['sku'])
                print(f'  [{done}/{total}] {pid}/v_{sid[:20]}   ', end='\r', flush=True)
                v['image'] = _upload(
                    v['image'], f'europolo/{pid}/v_{sid}')

    print(f'  Migration done: {done} images processed{" " * 30}')
    _save_cache()
    print(f'Cache saved: {CACHE_OUT}')
    print()

# ── Write product-data.json ────────────────────────────────────────────────────
with open(JSON_OUT, 'w', encoding='utf-8') as f:
    json.dump({'products': output}, f, ensure_ascii=False, indent=2)
print(f'Written: {JSON_OUT}')

# ── Write product-data-embed.js ────────────────────────────────────────────────
with open(JS_OUT, 'w', encoding='utf-8') as f:
    f.write('/* Auto-generated from data/product-data.json — do not edit manually */\n')
    f.write('window.EP_PRODUCTS = ')
    json.dump(output, f, ensure_ascii=False, separators=(',', ':'))
    f.write(';\n')
print(f'Written: {JS_OUT}')
print()
print('Done.')
